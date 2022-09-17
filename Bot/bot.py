from cgi import test
from xml.etree.ElementTree import tostring
from docx import Document
import openpyxl

workbook = openpyxl.load_workbook('trocar.xlsx')
sh = workbook.active

workbookCamaristas = openpyxl.load_workbook('Camaristas.xlsx')
shCam = workbookCamaristas.active

with open('texto_Feminino.txt', encoding='utf-8') as textoF:
    stringF = textoF.readlines()

with open('texto_Masculino.txt', encoding='utf-8') as textoM:
    stringM = textoM.readlines()

string = stringF
document = Document()

for i in range(2, shCam.max_row+1):

    celulaNome = shCam.cell(row=i, column=1)
    celulaSexo = shCam.cell(row=i, column=2)

    print(celulaNome.value + " - " + celulaSexo.value)

    for linha in range(0, len(string)):
        if celulaSexo.value == "Masculino":
            string[linha] = stringM[linha].replace("Nome do camarista", celulaNome.value)

        if celulaSexo.value == "Feminino":
            string[linha] = stringF[linha].replace("Nome do camarista", celulaNome.value)

        document.add_paragraph(string[linha])

    document.add_page_break()

document.save('teste.docx')
textoM.close()
textoF.close()